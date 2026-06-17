import hashlib
import io
import openpyxl


def compute_sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def xlsx_to_text(data: bytes) -> str:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise ValueError(f"Невозможно прочитать xlsx: {exc}") from exc

    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]

            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))

        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    if not parts:
        raise ValueError("В xlsx нет данных")

    return "\n\n".join(parts)
